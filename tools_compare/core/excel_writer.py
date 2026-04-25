"""导出 Excel: 汇总表 + 明细表。"""
from __future__ import annotations
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .diff_engine import AirportDiff


_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
_BODY_FONT = Font(name="微软雅黑", size=10)
_BORDER = Border(*(Side(style="thin", color="9CB6CF"),) * 4)
_WRAP = Alignment(wrap_text=True, vertical="center", horizontal="left")
_TYPE_FILL = {
    "新增": PatternFill("solid", fgColor="E2F0D9"),
    "删除": PatternFill("solid", fgColor="FCE4D6"),
    "修改": PatternFill("solid", fgColor="FFF2CC"),
}


def _style_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _BORDER


def _autosize(ws, widths: dict[int, int]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_diff_excel(diffs: Iterable[AirportDiff], out_path: str | Path) -> Path:
    out_path = Path(out_path)
    wb = Workbook()

    # ── 汇总表 ──
    ws = wb.active
    ws.title = "汇总"
    _style_header(ws, ["机场", "新旧数据对比结果", "差异条数", "备注"])
    diffs = list(diffs)
    for i, d in enumerate(diffs, 2):
        ws.cell(row=i, column=1, value=d.icao)
        ws.cell(row=i, column=2, value=d.summary_text)
        ws.cell(row=i, column=3, value=len(d.items))
        ws.cell(row=i, column=4, value=d.error or "")
        for col in range(1, 5):
            cell = ws.cell(row=i, column=col)
            cell.font = _BODY_FONT
            cell.alignment = _WRAP
            cell.border = _BORDER
    _autosize(ws, {1: 10, 2: 80, 3: 10, 4: 30})
    ws.freeze_panes = "A2"

    # ── 明细表 ──
    ws2 = wb.create_sheet("明细")
    _style_header(ws2, ["机场", "小节", "变更类型", "旧内容", "新内容"])
    row = 2
    for d in diffs:
        for it in d.items:
            ws2.cell(row=row, column=1, value=d.icao)
            ws2.cell(row=row, column=2, value=it.section)
            ws2.cell(row=row, column=3, value=it.change_type)
            ws2.cell(row=row, column=4, value=it.old_text)
            ws2.cell(row=row, column=5, value=it.new_text)
            for col in range(1, 6):
                cell = ws2.cell(row=row, column=col)
                cell.font = _BODY_FONT
                cell.alignment = _WRAP
                cell.border = _BORDER
            fill = _TYPE_FILL.get(it.change_type)
            if fill:
                ws2.cell(row=row, column=3).fill = fill
            row += 1
    _autosize(ws2, {1: 10, 2: 30, 3: 10, 4: 60, 5: 60})
    ws2.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
